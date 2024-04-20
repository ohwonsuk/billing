import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, time
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from dateutil.parser import parse #텍스트날짜를 날짜 형식으로 변경
import warnings
import re
import os

warnings.filterwarnings(action='ignore')  #경로 무시, 다시 적용시 default

st.title('스마트링크 청구내역서 관리')
st.header('월간청구내역서 생성')
# st.write(os.path.realpath(__file__))



# 차량번호 cleansing 함수
def carnoclean(car):
    carno = re.sub('\([^)]*\)', '', car)    
    if '(삭제)_고객사변경' in str(carno):
        return carno.replace('(삭제)_고객사변경', '')
    elif '_고객사변경' in str(carno):
        return carno.replace('_고객사변경', '')
    elif '__고객사변경' in str(carno):
        return carno.replace('__고객사변경', '')
    elif '(삭제)' in str(carno):
        return carno.replace('(삭제)', '')
    elif '(반납)' in str(carno):
        return carno.replace('(반납)', '')
    elif '(교체)' in str(carno):
        return carno.replace('(교체)', '')
    elif '(진행불가)' in str(carno):
        return carno.replace('(진행불가)', '')
    elif '(임시)' in str(carno):
        return carno.replace('(임시)', '')
    elif '(회수)' in str(carno):
        return carno.replace('(회수)', '')
    elif '(기존)' in str(carno):
        return carno.replace('(기존)', '')
    elif '(ㅅㅈ)' in str(carno):
        return carno.replace('(ㅅㅈ)', '')
    elif '(테스트)' in str(carno):
        return carno.replace('(테스트)', '')
    elif '__' in str(carno):
        return carno.replace('_', '')
    elif '_' in str(carno):
        return carno.replace('_', '')
    
    return carno

# 서비스명 조정 함수
def service_name(service):  
    name = re.sub('\([^)]*\)', '', service[0])
    if name != '-':       
        return name.replace('-', '')
    else:
        if service[1] == 'keybox':
            return '카셰어링프리미엄'
        return '차량운행관리'

# 이용료 누락 차량에 이용료 추가
def no_fare(service, price1, price2):
    # print(service[0], service[1])
    if (service[1] == 0.0) or (pd.isnull(service[1])):
        # print(service[1])
        if service[0] == '차량운행관리':
            # print(price1)
            return price1
        else:
            return price2
    return service[1]

# 청구기준일 이후 신규장착 차량 서비스시작일 변경
def service_start(service_date, start_date, data_start_date):
    if (pd.to_datetime(service_date) > start_date):
        return pd.to_datetime(service_date)
    elif (pd.to_datetime(service_date) > pd.to_datetime(data_start_date)):
        return pd.to_datetime(service_date)    
    return start_date

@st.cache_data
def sims():
    sims_raw=pd.read_excel("billing_car.xlsx")
    sims_raw['서비스1'] = sims_raw['서비스1'].fillna('-')
    sims_raw['equipnam2'] = sims_raw['equipnam2'].fillna('-')
    sims_raw['차량번호(clean)'] = sims_raw['차량번호'].apply(carnoclean)
    sims_raw['서비스명'] = sims_raw[['서비스1', 'equipnam2']].apply(service_name, axis=1)
    sims_raw['장착일'] = sims_raw['장착일'].apply(pd.to_datetime)
    return sims_raw

# 단가 일할 계산 (신규장착 - 장착일이 이용시작 이후 발생)
def price_cal(item, start_date, full_day, full_pday, price1, vat_include):
    if item[0] == price1:
        if pd.to_datetime(item[2]) > start_date:
            return round((item[0] / full_day * item[1]), -1)
        else:
            return -round((item[0] / full_pday * item[1]), -1)
    elif item[1] == full_day:
        if pd.to_datetime(item[2]) > start_date:
            return (item[0] / full_day * item[1])
        else:
            return -((item[0] / full_pday * item[1]))
    else:
        if pd.to_datetime(item[2]) > start_date:
            return round((item[0] / full_day * item[1]), -1)
        else:
            return -round((item[0] / full_pday * item[1]), -1)


def write_dataframe_to_excel(df, row, start_row, start_col):
    rows = list(dataframe_to_rows(df, index=False, header=False))
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            cell = wb['이용료'].cell(row = r_idx, column = c_idx, value = value)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')


customer_file = st.file_uploader('#### 청구고객사 엑셀파일을 업로드하세요 ####')
if customer_file is not None:
    customer_raw=pd.read_excel(customer_file)
    with st.expander("청구대상 고객사"):
        st.dataframe(customer_raw)
    # 청구양식 사용 고객사 리스트 추출하기
    customer_raw = customer_raw.drop(['순번'],axis=1)
    customer_raw.loc[customer_raw['주유'].str.contains('Y', na=False), '카드'] = 'Y'
    customer_raw.loc[customer_raw['하이패스'].str.contains('Y', na=False), '카드'] = 'Y'
    customer_raw['카드'] = customer_raw['카드'].fillna('N')
    customer_raw['요청기준일'] = customer_raw['요청기준일'].fillna('-')
    customer_list = customer_raw[customer_raw['계좌번호'].notnull() & (customer_raw['청구고객사'] != 'Y') ]
    customer_list['CMS고객사명'] = customer_list['CMS고객사명'].str.replace('㈜','(주)')
    customer_list['CMS고객사명'] = customer_list['CMS고객사명'].str.replace(' ','')
    customer_count = customer_list['법인명'].count()
    date_selection = customer_list['요청기준일'].unique()
    # customer_name_list = customer_list['CMS고객사명'].sort_values()

    st.write('직접청구 및 청구계좌번호 존재하는 고객사 대상')
    with st.expander("청구이력서 생성 고객사"):
        st.dataframe(customer_list)

    # customer=pd.read_excel("customer_name.xlsx")
    st.write('고객사 수 ', customer_count, ' 개사')

    # 청구기준일, 해당월 종료일, 해당월 날짜기간
    start_date = st.date_input('##### 청구기준일자 입력 #####', value=None)
    bill_date = st.selectbox('##### 청구서 작성일 입력 #####', (date_selection), index=None)
    # data_end_date = st.date_input('##### 청구서 작성일 입력 #####', datetime.now())

        # 청구서 작성일 기준 대상고객사 리스트 추출
    customer_filter = customer_list.loc[customer_list['요청기준일'] == bill_date]
    customer_name_list = customer_filter['CMS고객사명'].sort_values()
    print('custoemr filter-',customer_name_list)
    # print('청구서작성일', bill_date)
    if len(customer_name_list) == 0 :
        customer_filter = customer_list.loc[customer_list['요청기준일'] == '-']
        customer_name_list = customer_filter['CMS고객사명'].sort_values()

    st.write('청구기준일:', start_date)
    if bill_date is not None:
        bill_date_number = int(datetime.now().strftime('%d')) if  bill_date == '-' else int(bill_date.replace('일', '')) - 1 
        data_end_date = start_date + timedelta(days=bill_date_number)
        end_date = (start_date + relativedelta(months=1)- timedelta(days=1)).strftime('%Y-%m-%d')
        # full_day
        full_day = (datetime.strptime(end_date, '%Y-%m-%d').date() - start_date).days + 1

        #청구내역서 작성일기준 전월 날짜 계산
        data_start_date = (data_end_date - relativedelta(months=1) + timedelta(days=1)).strftime('%Y-%m-%d')
        year_pmonth = (data_end_date - relativedelta(months=1) + timedelta(days=1)).strftime("%Y%m")

        # 전달 말일 계산
        start_pdate = (start_date - relativedelta(months=1)).strftime('%Y-%m-%d')
        end_pdate = (start_date - timedelta(days=1)).strftime('%Y-%m-%d')
        full_pday = (datetime.strptime(end_pdate, '%Y-%m-%d').date() - datetime.strptime(start_pdate, '%Y-%m-%d').date()).days + 1

        #청구월, 연도, 연월 계산
        month = start_date.month
        year = start_date.year
        year_tmonth = start_date.strftime("%Y%m")

        st.write('청구서작성일:', data_end_date)

        select = st.selectbox(
            '청구내역서 개별 또는 전체 생성 선택하세요',
            ('개별', '전체'),index=None, placeholder='방법을 선택하기')

        if select == '개별':
            customer_name = st.selectbox(
                '고객사를 선택하세요',
                (customer_name_list))

            customer_bill = customer_list[customer_list['CMS고객사명'] == customer_name]
            customer_bill_name = customer_bill.iloc[0,1]  #청구서용 법인명 불러오기
            customer_code = customer_bill.iloc[0,10]    #사업자번호 불러오기
            customer_account = customer_bill.iloc[0,18] #계좌번호 불러오기
            card_use = customer_bill.iloc[0,31]   #하나카드 사용여부
            bill_month = customer_bill.iloc[0,27] #당월, 전월 구분
            card_percent = customer_bill.iloc[0,25] #카드 수수료 구분
            vat_include = customer_bill.iloc[0,29] #부가세 포함여부 (포함 = 'Y')
            service1 = customer_bill.iloc[0,19] #서비스명1 불러오기
            service2 = customer_bill.iloc[0,21] #서비스명2 불러오기
            print('서비스명-',service1)
            # 카셰어링베이직 서비스의 경우 양식의 서비스종류 명칭 변경
            if (service1 == '카셰어링베이직') or (service2 == '카셰어링베이직'):
                basic_service = 'Y'
            else: 
                basic_service = 'N'
                    # 고객사 서비스별 단가 매칭 
            company_data = list(dataframe_to_rows(customer_bill, index=False, header=False))
            price1 = company_data[0][20]  # 차량운행관리
            price2 = company_data[0][22]  # 카셰어링프리미엄
            print('이용료',company_data)

            st.write('청구고객사명:',customer_bill_name, ', 사업자등록번호:',customer_code, ', 계좌번호:',customer_account, ', 청구기준:', bill_month, ', 하나카드사용:', card_use, ', 카드수수료:', card_percent)

            if st.button('청구대상 차량 불러오기'):
                sims_raw = sims()
                with st.expander("청구대상 차량리스트"):
                    st.dataframe(sims_raw)
                sims_customer = sims_raw.loc[(sims_raw['고객명'] == customer_name) & (sims_raw['장착일'] <= end_date)]
                columns = ['차량번호(clean)', '차종', '서비스명', '단가1', '장착일']
                customer_car = sims_customer[columns]
            # 청구 고객사 차량 리스트 추출
            # if st.button('고객사 차량 불러오기'):
            # sims_customer = sims_raw.loc[sims_raw['고객명'] == customer_name]
                st.write(customer_name, ' 차량리스트')
                st.dataframe(sims_customer)
                st.write('청구내역서 항목')
                st.dataframe(customer_car)

                #CMS 데이터 기준 서비스 미사용 차량 조회 (월별로 사전에 생성)
                cms_tmonth = pd.read_excel(f"cms_off_list_({year_tmonth}).xlsx")
                cms_pmonth = pd.read_excel(f"cms_off_list_({year_pmonth}).xlsx")

                # cms 고개사 명칭 정리
                cms_tmonth['고객사'] = cms_tmonth['고객사'].str.replace('㈜','(주)')
                cms_tmonth['고객사'] = cms_tmonth['고객사'].str.replace(' ','')
                cms_pmonth['고객사'] = cms_pmonth['고객사'].str.replace('㈜','(주)')
                cms_pmonth['고객사'] = cms_pmonth['고객사'].str.replace(' ','')

                # 날짜 형식 변황
                car_toff = cms_tmonth.loc[cms_tmonth['고객사'] == customer_name]
                car_toff['종료일'] = car_toff['종료일자'].dt.strftime("%Y-%m-%d")
                car_poff = cms_pmonth.loc[cms_pmonth['고객사'] == customer_name]
                car_poff['종료일'] = car_poff['종료일자'].dt.strftime("%Y-%m-%d")

                #청구 고객사 정보와 차량 매칭
                car_toff_merge = pd.merge(car_toff, customer_bill, left_on='고객사', right_on='CMS고객사명', how='left')
                car_poff_merge = pd.merge(car_poff, customer_bill, left_on='고객사', right_on='CMS고객사명', how='left')
                columns = ['차량번호(clean)','모델','서비스명1', '단가1', '종료일']
                car_toff_list = car_toff_merge[columns]
                car_poff_list = car_poff_merge[columns]

                #청구 기준일 기준 1개월 이전의 서비스 종료차량 필터링
                car_poff_filter = car_poff_list.loc[pd.to_datetime(car_poff_list['종료일']) >= data_start_date]
                car_toff_filter = car_toff_list.loc[pd.to_datetime(car_toff_list['종료일']) <= pd.to_datetime(data_end_date)]
                car_off_list = pd.concat([car_poff_filter, car_toff_filter])
                #종료 차량 조회
                st.write('##### 종료차량 조회 #####')
                st.dataframe(car_off_list)

                # 청구 데이터 생성
                customer_car['단말기상태'] = '이용'
                customer_car['계약기간'] = '-'

                # s_date = pd.Timestamp(start_date)
                # start_date_str = start_date.strftime('%Y-%m-%d')
                # customer_car['이용시작'] = customer_car['장착일'].apply(service_start, s_date =s_date)
                # customer_car['이용시작'] = customer_car['이용시작'].dt.strftime("%Y-%m-%d")
                customer_car['이용시작'] = customer_car['장착일'].apply(service_start, args=(pd.to_datetime(start_date), pd.to_datetime(data_start_date)))

                # 말일 날짜 계산 (1개월 더해서 1일 빼기)
                customer_car['이용종료'] = end_date
                customer_car.reset_index(drop=True, inplace=True)
                # 청구내역서 작성일 이전 이용시작 차량만 추출
                customer_car_filter = customer_car.loc[customer_car['이용시작'] <= pd.to_datetime(data_end_date)]

                # 이용료 누락 차량 단가 채우기
                customer_car_filter['단가1'] = customer_car_filter[['서비스명', '단가1']].apply(no_fare, args=(price1, price2), axis=1)
                customer_car_filter['이용시작'] = customer_car_filter['이용시작'].dt.strftime("%Y-%m-%d")
                # st.dataframe(customer_car)
                # 청구대상 차량대수
                number_of_list = customer_car['차량번호(clean)'].count()
                # 종료차량 청구대상에 추가하기
                append_data = list(dataframe_to_rows(car_off_list, index=False, header=False))
                for r_idx, row in enumerate(append_data):
                            if pd.to_datetime(row[4]) >  pd.to_datetime(start_date):
                                customer_car_filter.loc[number_of_list + r_idx] = [row[0], row[1], row[2], row[3], row[4], '반납', '-', start_date.strftime('%Y-%m-%d'), row[4]]
                            else:
                                customer_car_filter.loc[number_of_list + r_idx] = [row[0], row[1], row[2], row[3], row[4], '반납', '-', row[4], end_pdate]

                #청구대상 기산 산정
                customer_car_filter['start']=pd.to_datetime(customer_car_filter['이용시작'])
                customer_car_filter['end']=pd.to_datetime(customer_car_filter['이용종료'])
                customer_car_filter['gap']= customer_car_filter['end'] - customer_car_filter['start']
                customer_car_filter['gap'] = customer_car_filter['gap'].dt.days + 1

                # 단가 일할 계산 (신규장착 - 장착일이 이용시작 이후 발생)
                customer_car_filter['공급가액'] = customer_car_filter[['단가1', 'gap', 'end']].apply(price_cal, args=[pd.to_datetime(start_date), full_day, full_pday, price1, vat_include], axis=1)
                #청구양식 만들기
                columns = ['차량번호(clean)', '차종','서비스명','단말기상태','계약기간','이용시작', '이용종료', '공급가액']
                car_list = customer_car_filter[columns]
                car_list.columns = ['차량번호', '차종', '서비스구분', '단말기상태','계약기간','이용시작', '이용종료', '공급가액']
                car_sort = car_list.sort_values(by=['단말기상태' ,'공급가액'])
                # 청구대상 대수 확정
                row = car_list['차량번호'].count()
                card_filename = 'card-form1.xlsx' if card_percent == 0.01 else 'card-form2.xlsx'

                st.write(customer_bill_name," 청구리스트")
                st.dataframe(car_sort) 
                # st.data_editor(car_sort, num_rows="dynamic")
                print('계좌번호',customer_account)
                wb = (load_workbook('basic-form.xlsx') if card_use != 'Y' else load_workbook(card_filename))

                # 이용료 내역 만들기
                ws2 = wb['이용료']
                ws2['B2'].value = f'{year}년 {month}월' 
                ws2['B4'].value = customer_code      #사업자등록번호
                ws2['C4'].value = customer_bill_name  #고객사명
                # 스타일 지정
                border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                                                right=openpyxl.styles.Side(style='thin'), 
                                                top=openpyxl.styles.Side(style='thin'), 
                                                bottom=openpyxl.styles.Side(style='thin'))
                font = openpyxl.styles.Font(name='맑은 고딕', size=9)

                write_dataframe_to_excel(car_sort, row-1, 6, 2)
                for i in range(6, 6+row):
                    ws2['I'+str(i)].alignment = Alignment(horizontal='right', vertical='center')
                    ws2['J'+str(i)] = "=round(I"+str(i)+"*0.1,0)"
                    ws2['J'+str(i)].number_format = '#,##0'
                    ws2['J'+str(i)].border = border
                    ws2['J'+str(i)].font = font
                    ws2['K'+str(i)] = "=I"+str(i)+"+J"+str(i)
                    ws2['K'+str(i)].number_format = '#,##0'
                    ws2['K'+str(i)].border = border
                    ws2['K'+str(i)].font = font
                    ws2['L'+str(i)].border = border
                    ws2['M'+str(i)].border = border
                    ws2['I'+str(i)].number_format = '#,##0'

                # 카드상세 내역 만들기
                if card_use == 'Y':
                    ws3 = wb['카드상세내역']
                    if month != 1:
                        ws3['B2'].value = f'{year}년 {month-1}월' 
                    else:
                        ws3['B2'].value = f'{year-1}년 {month+11}월' 

                # 청구서 표지 만들기
                #   st.write('청구표지 만들기')
                ws1 = wb['청구서']
                
                # print(ws1['G15'].value)
                ws1['B4'].value = customer_bill_name  #고객사명
                ws1['B6'].value = f'{month}월 이용대금 청구서'
                ws1['O1'].value = data_end_date.strftime('%Y-%m-%d')  #청구서작성일자

                if card_use == 'N':
                    ws1['I25'].value = customer_account   #계좌번호
                else:
                    ws1['I33'].value = customer_account   #계좌번호

                if basic_service == 'Y':
                    print('카세어링베이직',basic_service)
                    ws1['B16'].value = '카셰어링베이직'
                print(ws1['I25'].value)
                wb.save(f'/Users/wonsuk/Documents/streamlit/billing/output/{customer_bill_name}_{month}월_스마트링크내역서.xlsx')
                st.write('청구내역서 생성완료')
        else:
            if select == '전체':
                print('고객사리스트', type(customer_filter), customer_filter)
                if st.button('청구내역서 만들기'):
                    # customers = customer_filter.values.tolist()
                    for name in customer_name_list:
                        print('고객사명',name)
                        customer_bill = customer_list[customer_list['CMS고객사명'] == name]
                        print('bill-',customer_bill)
                        customer_bill_name = customer_bill.iloc[0,1]  #청구서용 법인명 불러오기
                        customer_code = customer_bill.iloc[0,10]    #사업자번호 불러오기
                        customer_account = customer_bill.iloc[0,18] #계좌번호 불러오기
                        card_use = customer_bill.iloc[0,31]   #하나카드 사용여부
                        bill_month = customer_bill.iloc[0,27]  #당월, 전월 기준
                        card_percent = customer_bill.iloc[0,25] #카드수수료
                        vat_include = customer_bill.iloc[0,29] #부가세 포함여부 (포함 = 'Y')
                        service1 = customer_bill.iloc[0,19] #서비스명1 불러오기
                        service2 = customer_bill.iloc[0,21] #서비스명2 불러오기
                        print('서비스명-',service1)
                        # 카셰어링베이직 서비스의 경우 양식의 서비스종류 명칭 변경
                        if (service1 == '카셰어링베이직') or (service2 == '카셰어링베이직'):
                            basic_service = 'Y'
                        else: 
                            basic_service = 'N'
                        company_data = list(dataframe_to_rows(customer_bill, index=False, header=False))
                        price1 = company_data[0][20]  # 차량운행관리
                        price2 = company_data[0][22]  # 카셰어링프리미엄
                        end_date = (start_date + relativedelta(months=1)- timedelta(days=1)).strftime('%Y-%m-%d')
                        # full_day
                        # full_day = (datetime.strptime(end_date, '%Y-%m-%d').date() - start_date).days
                        #청구월, 연도, 연월 계산
                        # month = start_date.month
                        # year = start_date.year
                        # year_month = start_date.strftime("%Y%m")
                        sims_raw = sims()
                        sims_customer = sims_raw.loc[(sims_raw['고객명'] == name) & (sims_raw['장착일'] <= end_date)]
                        columns = ['차량번호(clean)', '차종', '서비스명', '단가1', '장착일']
                        customer_car = sims_customer[columns]

                        #CMS 데이터 기준 서비스 미사용 차량 조회 (월별로 사전에 생성)
                        cms_tmonth = pd.read_excel(f"cms_off_list_({year_tmonth}).xlsx")
                        cms_pmonth = pd.read_excel(f"cms_off_list_({year_pmonth}).xlsx")

                        # cms 고개사 명칭 정리
                        cms_tmonth['고객사'] = cms_tmonth['고객사'].str.replace('㈜','(주)')
                        cms_tmonth['고객사'] = cms_tmonth['고객사'].str.replace(' ','')
                        cms_pmonth['고객사'] = cms_pmonth['고객사'].str.replace('㈜','(주)')
                        cms_pmonth['고객사'] = cms_pmonth['고객사'].str.replace(' ','')

                        # 날짜 형식 변황
                        car_toff = cms_tmonth.loc[cms_tmonth['고객사'] == name]
                        car_toff['종료일'] = car_toff['종료일자'].dt.strftime("%Y-%m-%d")
                        car_poff = cms_pmonth.loc[cms_pmonth['고객사'] == name]
                        car_poff['종료일'] = car_poff['종료일자'].dt.strftime("%Y-%m-%d")

                        #청구 고객사 정보와 차량 매칭
                        car_toff_merge = pd.merge(car_toff, customer_bill, left_on='고객사', right_on='CMS고객사명', how='left')
                        car_poff_merge = pd.merge(car_poff, customer_bill, left_on='고객사', right_on='CMS고객사명', how='left')
                        columns = ['차량번호(clean)','모델','서비스명1', '단가1', '종료일']
                        car_toff_list = car_toff_merge[columns]
                        car_poff_list = car_poff_merge[columns]

                        #청구 기준일 기준 1개월 이전의 서비스 종료차량 필터링
                        car_poff_filter = car_poff_list.loc[pd.to_datetime(car_poff_list['종료일']) >= data_start_date]
                        car_toff_filter = car_toff_list.loc[pd.to_datetime(car_toff_list['종료일']) <= pd.to_datetime(data_end_date)]
                        car_off_list = pd.concat([car_poff_filter, car_toff_filter])
                        #종료 차량 조회
                        # st.write('##### 종료차량 조회 #####')
                        # st.dataframe(car_off_list)

                        # 청구 데이터 생성
                        customer_car['단말기상태'] = '이용'
                        customer_car['계약기간'] = '-'

                        # s_date = pd.Timestamp(start_date)
                        # start_date_str = start_date.strftime('%Y-%m-%d')
                        # customer_car['이용시작'] = customer_car['장착일'].apply(service_start, s_date =s_date)
                        # customer_car['이용시작'] = customer_car['이용시작'].dt.strftime("%Y-%m-%d")
                        customer_car['이용시작'] = customer_car['장착일'].apply(service_start, args=(pd.to_datetime(start_date), pd.to_datetime(data_start_date)))

                        # 말일 날짜 계산 (1개월 더해서 1일 빼기)
                        customer_car['이용종료'] = end_date
                        customer_car.reset_index(drop=True, inplace=True)
                        # 청구내역서 작성일 이전 이용시작 차량만 추출
                        customer_car_filter = customer_car.loc[customer_car['이용시작'] <= pd.to_datetime(data_end_date)]

                        # 이용료 누락 차량 단가 채우기
                        customer_car_filter['단가1'] = customer_car_filter[['서비스명', '단가1']].apply(no_fare, args=(price1, price2), axis=1)
                        customer_car_filter['이용시작'] = customer_car_filter['이용시작'].dt.strftime("%Y-%m-%d")
                        # st.dataframe(customer_car)
                        # 청구대상 차량대수
                        number_of_list = customer_car['차량번호(clean)'].count()
                        # 종료차량 청구대상에 추가하기
                        append_data = list(dataframe_to_rows(car_off_list, index=False, header=False))
                        for r_idx, row in enumerate(append_data):
                                    if pd.to_datetime(row[4]) > pd.to_datetime(start_date):
                                        customer_car_filter.loc[number_of_list + r_idx] = [row[0], row[1], row[2], row[3], row[4], '반납', '-', start_date.strftime('%Y-%m-%d'), row[4]]
                                    else:
                                        customer_car_filter.loc[number_of_list + r_idx] = [row[0], row[1], row[2], row[3], row[4], '반납', '-', row[4], end_pdate]

                        #청구대상 기산 산정
                        customer_car_filter['start']=pd.to_datetime(customer_car_filter['이용시작'])
                        customer_car_filter['end']=pd.to_datetime(customer_car_filter['이용종료'])
                        customer_car_filter['gap']= customer_car_filter['end'] - customer_car_filter['start']
                        customer_car_filter['gap'] = customer_car_filter['gap'].dt.days + 1

                        # 단가 일할 계산 (신규장착 - 장착일이 이용시작 이후 발생)
                        customer_car_filter['공급가액'] = customer_car_filter[['단가1', 'gap', 'end']].apply(price_cal, args=[pd.to_datetime(start_date), full_day, full_pday, price1, vat_include], axis=1)
                        #청구양식 만들기
                        columns = ['차량번호(clean)', '차종','서비스명','단말기상태','계약기간','이용시작', '이용종료', '공급가액']
                        car_list = customer_car_filter[columns]
                        car_list.columns = ['차량번호', '차종', '서비스구분', '단말기상태','계약기간','이용시작', '이용종료', '공급가액']
                        car_sort = car_list.sort_values(by=['단말기상태' ,'공급가액'])
                        # 청구대상 대수 확정
                        row = car_list['차량번호'].count()
                        card_filename = 'card-form1.xlsx' if card_percent == 0.01 else 'card-form2.xlsx'

                        # st.write(customer_name," 청구리스트")
                        # st.dataframe(car_sort) 
                        # st.data_editor(car_sort, num_rows="dynamic")
                        wb = (load_workbook('basic-form.xlsx') if card_use != 'Y' else load_workbook(card_filename))

                        # 이용료 내역 만들기
                        ws2 = wb['이용료']
                        ws2['B2'].value = f'{year}년 {month}월' 
                        ws2['B4'].value = customer_code      #사업자등록번호
                        ws2['C4'].value = customer_bill_name  #고객사명

                        # 스타일 지정
                        border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                                                        right=openpyxl.styles.Side(style='thin'), 
                                                        top=openpyxl.styles.Side(style='thin'), 
                                                        bottom=openpyxl.styles.Side(style='thin'))
                        font = openpyxl.styles.Font(name='맑은 고딕', size=9)
                        write_dataframe_to_excel(car_sort, row-1, 6, 2)
                        for i in range(6, 6+row):
                            ws2['I'+str(i)].alignment = Alignment(horizontal='right', vertical='center')
                            ws2['J'+str(i)] = "=round(I"+str(i)+"*0.1,0)"
                            ws2['J'+str(i)].number_format = '#,##0'
                            ws2['J'+str(i)].border = border
                            ws2['J'+str(i)].font = font
                            ws2['K'+str(i)] = "=I"+str(i)+"+J"+str(i)
                            ws2['K'+str(i)].number_format = '#,##0'
                            ws2['K'+str(i)].border = border
                            ws2['K'+str(i)].font = font
                            ws2['L'+str(i)].border = border
                            ws2['M'+str(i)].border = border
                            ws2['I'+str(i)].number_format = '#,##0'

                        # 카드상세 내역 만들기
                        if card_use == 'Y':
                            ws3 = wb['카드상세내역']
                            if month != 1:
                                ws3['B2'].value = f'{year}년 {month-1}월' 
                            else:
                                ws3['B2'].value = f'{year-1}년 {month+11}월' 
                        # 청구서 표지 만들기
                        #   st.write('청구표지 만들기')
                        ws1 = wb['청구서']
                        # print(ws1['E15'].value)
                        # print(ws1['G15'].value)
                        # print(ws1['I25'].value)
                        ws1['B4'].value = customer_bill_name  #고객사명
                        ws1['B6'].value = f'{month}월 이용대금 청구서'
                        ws1['O1'].value = data_end_date.strftime('%Y-%m-%d')   #청구서작성일자
                        if basic_service == 'Y':
                            print('카세어링베이직',basic_service)
                            ws1['B16'].value = '카셰어링베이직'

                        if card_use == 'N':
                            ws1['I25'].value = customer_account   #계좌번호
                        else:
                            ws1['I33'].value = customer_account   #계좌번호


                        wb.save(f'/Users/wonsuk/Documents/streamlit/billing/output/{customer_bill_name}_{month}월_스마트링크내역서.xlsx')
                        st.write(customer_bill_name,'-청구내역서 생성완료')


else:
    st.warning('엑셀파일을 업로드 하세요')